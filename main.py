import requests
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from bs4 import BeautifulSoup
from openpyxl import load_workbook

class Sheets:

    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name('scr.json', scope)
    client = gspread.authorize(creds)
    sheet = client.open('markastok').sheet1

    def __init__(self, urllink, name, off, disc, sale, avail):

        self.urllink = urllink
        self.name = name
        self.off = off
        self. disc = disc
        self.sale = sale
        self.avail = avail

    def addrow(self):

        rownum = 2
        Sheets.sheet.insert_row([self.urllink, self.name, self.disc, self.off, self.sale, self.avail], rownum)
        rownum = rownum+1

class Scraper(Sheets):

    def __init__(self, urllink, name, off, disc, sale, avail):
        super().__init__(urllink, name, off, disc, sale, avail)

    def scrape(self):

        source = requests.get(self.urllink).text
        soup = BeautifulSoup(source, 'lxml')

        if soup.find('div', id='productInfo'):
            product_name = soup.find('div', id='productInfo').h1.text
            self.name = product_name

            if soup.find('span', class_="product-price"):

                offer = soup.find('div', class_="detay-indirim").text
                self.off = str(offer)
                product_price = soup.find('span', class_="currencyPrice discountedPrice").text
                self.disc = str(product_price)
                sale_price = soup.find('span', class_="product-price").text
                self.sale = str(sale_price)
                avl = len(soup.find_all('a', class_="col box-border"))
                pasAvl = len(soup.find_all('a', class_="col box-border passive"))
                percent = (avl/(avl+pasAvl))*100
                self.avail = percent

            else:

                self.off = '-'
                self.disc = '-'
                self.sale = '-'
                self.avail = 0

            super().addrow()


def Urls():

    wb = load_workbook('url.xlsx')
    ws = wb.active
    url_list = []
    for i in range(1, 1000):
        url_list.append('http://www.markastok.com' + str(ws['A' + str(i)].value))

    for l in range(len(url_list)-1):

        Scraper(url_list[l], 'name', 'price', 'offer', 'sale price', 0).scrape()

if __name__ == '__main__':

    print(':)')
    Urls()

